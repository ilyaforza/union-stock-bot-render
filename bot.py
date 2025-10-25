import os
import openpyxl
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler
import ftplib
import io
import logging
from datetime import datetime, timedelta
import asyncio
import re
import json
import pytz
import requests
from sqlalchemy import create_engine, Column, Integer, String, Boolean, DateTime, Text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from dotenv import load_dotenv

# Загрузка переменных окружения
load_dotenv()

# 🔐 НАСТРОЙКИ ИЗ ПЕРЕМЕННЫХ СРЕДЫ
BOT_TOKEN = os.environ.get('BOT_TOKEN')
ADMIN_ID = int(os.environ.get('ADMIN_ID', 275066977))

# FTP настройки
FTP_HOST = os.environ.get('FTP_HOST')
FTP_PORT = int(os.environ.get('FTP_PORT', 21))
FTP_USERNAME = os.environ.get('FTP_USERNAME')
FTP_PASSWORD = os.environ.get('FTP_PASSWORD')
FTP_PATH = os.environ.get('FTP_PATH', '/')
FTP_FILENAME = os.environ.get('FTP_FILENAME', "Ostatki dlya bota (XLSX).xlsx")

# Настройка PostgreSQL
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

# Локальный файл для резервного копирования
LOCAL_FILENAME = "Ostatki dlya bota (XLSX).xlsx"

# Московский часовой пояс
MOSCOW_TZ = pytz.timezone('Europe/Moscow')

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Инициализация базы данных PostgreSQL
Base = declarative_base()

class User(Base):
    __tablename__ = 'users'
    
    user_id = Column(Integer, primary_key=True)
    username = Column(String(100))
    first_name = Column(String(100))
    last_name = Column(String(100))
    is_blocked = Column(Boolean, default=False)
    is_approved = Column(Boolean, default=False)
    block_reason = Column(Text)
    block_until = Column(DateTime)
    request_count = Column(Integer, default=0)
    first_seen = Column(DateTime)
    last_seen = Column(DateTime)
    approval_requested = Column(DateTime)

class AdminLog(Base):
    __tablename__ = 'admin_logs'
    
    id = Column(Integer, primary_key=True, autoincrement=True)
    admin_id = Column(Integer)
    action = Column(String(200))
    target_user_id = Column(Integer)
    details = Column(Text)
    timestamp = Column(DateTime)

class BotData(Base):
    __tablename__ = 'bot_data'
    
    id = Column(Integer, primary_key=True)
    key = Column(String(100), unique=True)
    value = Column(Text)
    updated_at = Column(DateTime)

# Инициализация базы данных
def init_db():
    try:
        engine = create_engine(DATABASE_URL)
        Base.metadata.create_all(engine)
        logger.info("✅ База данных PostgreSQL инициализирована")
        return engine
    except Exception as e:
        logger.error(f"❌ Ошибка инициализации базы данных: {e}")
        # Создаем временную SQLite базу как запасной вариант
        fallback_db = create_engine('sqlite:///fallback.db')
        Base.metadata.create_all(fallback_db)
        logger.info("✅ Создана резервная SQLite база данных")
        return fallback_db

# Глобальный engine для базы данных
try:
    engine = init_db()
    Session = sessionmaker(bind=engine)
except Exception as e:
    logger.error(f"❌ Критическая ошибка базы данных: {e}")
    raise

# Функции для работы с базой данных
def get_user(user_id):
    session = Session()
    try:
        user = session.query(User).filter(User.user_id == user_id).first()
        return user
    finally:
        session.close()

def update_user(user_id, username, first_name, last_name):
    session = Session()
    try:
        user = get_user(user_id)
        now = datetime.now(MOSCOW_TZ)
        
        if user:
            user.username = username
            user.first_name = first_name
            user.last_name = last_name
            user.last_seen = now
            user.request_count += 1
        else:
            is_approved = True if user_id == ADMIN_ID else False
            user = User(
                user_id=user_id,
                username=username,
                first_name=first_name,
                last_name=last_name,
                first_seen=now,
                last_seen=now,
                request_count=1,
                approval_requested=now,
                is_approved=is_approved
            )
            session.add(user)
        
        session.commit()
    finally:
        session.close()

def approve_user(user_id):
    session = Session()
    try:
        user = session.query(User).filter(User.user_id == user_id).first()
        if user:
            user.is_approved = True
            session.commit()
    finally:
        session.close()

def block_user(user_id, reason="Не указана", block_until=None):
    session = Session()
    try:
        user = session.query(User).filter(User.user_id == user_id).first()
        if user:
            user.is_blocked = True
            user.block_reason = reason
            user.block_until = block_until
            session.commit()
    finally:
        session.close()

def unblock_user(user_id):
    session = Session()
    try:
        user = session.query(User).filter(User.user_id == user_id).first()
        if user:
            user.is_blocked = False
            user.block_reason = None
            user.block_until = None
            session.commit()
    finally:
        session.close()

def get_all_users():
    session = Session()
    try:
        users = session.query(User).order_by(User.last_seen.desc()).all()
        return users
    finally:
        session.close()

def get_pending_approvals():
    session = Session()
    try:
        users = session.query(User).filter(
            User.is_approved == False,
            User.is_blocked == False,
            User.user_id != ADMIN_ID
        ).order_by(User.first_seen.desc()).all()
        return users
    finally:
        session.close()

def log_admin_action(admin_id, action, target_user_id=None, details=None):
    session = Session()
    try:
        log = AdminLog(
            admin_id=admin_id,
            action=action,
            target_user_id=target_user_id,
            details=details,
            timestamp=datetime.now(MOSCOW_TZ)
        )
        session.add(log)
        session.commit()
    finally:
        session.close()

def get_bot_data(key, default=None):
    session = Session()
    try:
        data = session.query(BotData).filter(BotData.key == key).first()
        return data.value if data else default
    finally:
        session.close()

def set_bot_data(key, value):
    session = Session()
    try:
        data = session.query(BotData).filter(BotData.key == key).first()
        if data:
            data.value = value
            data.updated_at = datetime.now(MOSCOW_TZ)
        else:
            data = BotData(key=key, value=value, updated_at=datetime.now(MOSCOW_TZ))
            session.add(data)
        session.commit()
    finally:
        session.close()

# Проверка доступа пользователя
def is_user_allowed(user_id):
    if user_id == ADMIN_ID:
        return True
    
    user = get_user(user_id)
    if not user:
        return False
    
    if user.is_blocked:
        if user.block_until:
            if datetime.now(MOSCOW_TZ) > user.block_until:
                unblock_user(user_id)
                return user.is_approved
        return False
    
    return user.is_approved

class StockBot:
    def __init__(self):
        self.products = []
        self.shipment_dates = []
        self.last_update = None
        self.data_source = "Не загружено"
        self.file_modify_time = None
        self.auto_update_enabled = True
        self.last_auto_update = None
        
    def load_data(self):
        """Загрузка данных - сначала пробуем FTP, потом локальный файл"""
        # Пробуем загрузить с FTP
        if self.download_file_from_ftp():
            self.data_source = "FTP сервер"
            return True
        
        # Если FTP не сработал, пробуем локальный файл
        if self.load_local_file():
            self.data_source = "Локальный файл"
            return True
            
        return False
        
    def download_file_from_ftp(self):
        """Загрузка файла с FTP сервера"""
        try:
            ftp = ftplib.FTP()
            ftp.connect(FTP_HOST, FTP_PORT)
            ftp.login(FTP_USERNAME, FTP_PASSWORD)
            
            try:
                ftp.cwd(FTP_PATH)
            except:
                logger.warning(f"Не удалось перейти в папку {FTP_PATH}, пробуем корневую")
            
            # Получаем время модификации файла
            try:
                file_time = ftp.voidcmd(f"MDTM {FTP_FILENAME}")[4:].strip()
                utc_time = datetime.strptime(file_time, '%Y%m%d%H%M%S')
                self.file_modify_time = utc_time.replace(tzinfo=pytz.utc).astimezone(MOSCOW_TZ)
            except:
                logger.warning("Не удалось получить время модификации файла с FTP")
                self.file_modify_time = datetime.now(MOSCOW_TZ)
            
            # Загружаем файл в память
            file_data = io.BytesIO()
            ftp.retrbinary(f'RETR {FTP_FILENAME}', file_data.write)
            file_data.seek(0)
            
            ftp.quit()
            
            # Читаем Excel файл
            workbook = openpyxl.load_workbook(file_data)
            sheet = workbook['TDSheet']
            
            # Собираем даты поставок из строки 4 (G4-Z4)
            self.shipment_dates = []
            for col in range(7, 27):
                date_cell = sheet.cell(row=4, column=col).value
                if date_cell and self._parse_date(date_cell):
                    self.shipment_dates.append({
                        'column': col,
                        'date': self._parse_date(date_cell),
                        'display_date': str(date_cell).strip()
                    })
            
            self.products = []
            
            # Парсим данные начиная с 6-й строки
            for row in range(6, sheet.max_row + 1):
                product_name = sheet.cell(row=row, column=1).value
                
                if not product_name:
                    continue
                    
                product_name_str = str(product_name)
                skip_keywords = ['Остатки', 'Номенклатура', 'Итого', '2.SPC', '1.UNION', 
                               '2.Essence', '3.Art', '4.Creative', '4.Подложка', '5.Клей']
                
                if any(keyword in product_name_str for keyword in skip_keywords):
                    continue
                
                # Столбец E - В резерве
                reserve_value = sheet.cell(row=row, column=5).value
                # Столбец F - Доступно
                available_value = sheet.cell(row=row, column=6).value
                
                # Столбцы C и D - дополнительная информация
                info_c = sheet.cell(row=row, column=3).value
                info_d = sheet.cell(row=row, column=4).value
                additional_info = ""
                if info_c:
                    additional_info += str(info_c)
                if info_d:
                    if additional_info:
                        additional_info += " "
                    additional_info += str(info_d)
                
                # Собираем информацию о поставках для этого товара
                shipments = {}
                for date_info in self.shipment_dates:
                    col = date_info['column']
                    shipment_value = sheet.cell(row=row, column=col).value
                    if shipment_value and self._parse_value(shipment_value) > 0:
                        shipments[date_info['display_date']] = self._parse_value(shipment_value)
                
                product = {
                    'name': product_name_str,
                    'additional_info': additional_info,
                    'reserve': self._parse_value(reserve_value),
                    'available': self._parse_value(available_value),
                    'shipments': shipments
                }
                
                self.products.append(product)
            
            self.last_update = datetime.now(MOSCOW_TZ)
            logger.info(f"Файл успешно загружен с FTP. Найдено {len(self.products)} товаров и {len(self.shipment_dates)} дат поставок")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке файла с FTP: {e}")
            return False
    
    def load_local_file(self):
        """Загрузка локального файла"""
        try:
            if os.path.exists(LOCAL_FILENAME):
                file_stat = os.stat(LOCAL_FILENAME)
                utc_time = datetime.fromtimestamp(file_stat.st_mtime)
                self.file_modify_time = utc_time.replace(tzinfo=pytz.utc).astimezone(MOSCOW_TZ)
            else:
                self.file_modify_time = datetime.now(MOSCOW_TZ)
            
            workbook = openpyxl.load_workbook(LOCAL_FILENAME)
            sheet = workbook['TDSheet']
            
            self.shipment_dates = []
            for col in range(7, 27):
                date_cell = sheet.cell(row=4, column=col).value
                if date_cell and self._parse_date(date_cell):
                    self.shipment_dates.append({
                        'column': col,
                        'date': self._parse_date(date_cell),
                        'display_date': str(date_cell).strip()
                    })
            
            self.products = []
            
            for row in range(6, sheet.max_row + 1):
                product_name = sheet.cell(row=row, column=1).value
                
                if not product_name:
                    continue
                    
                product_name_str = str(product_name)
                skip_keywords = ['Остатки', 'Номенклатура', 'Итого', '2.SPC', '1.UNION', 
                               '2.Essence', '3.Art', '4.Creative', '4.Подложка', '5.Клей']
                
                if any(keyword in product_name_str for keyword in skip_keywords):
                    continue
                
                reserve_value = sheet.cell(row=row, column=5).value
                available_value = sheet.cell(row=row, column=6).value
                
                info_c = sheet.cell(row=row, column=3).value
                info_d = sheet.cell(row=row, column=4).value
                additional_info = ""
                if info_c:
                    additional_info += str(info_c)
                if info_d:
                    if additional_info:
                        additional_info += " "
                    additional_info += str(info_d)
                
                shipments = {}
                for date_info in self.shipment_dates:
                    col = date_info['column']
                    shipment_value = sheet.cell(row=row, column=col).value
                    if shipment_value and self._parse_value(shipment_value) > 0:
                        shipments[date_info['display_date']] = self._parse_value(shipment_value)
                
                product = {
                    'name': product_name_str,
                    'additional_info': additional_info,
                    'reserve': self._parse_value(reserve_value),
                    'available': self._parse_value(available_value),
                    'shipments': shipments
                }
                
                self.products.append(product)
            
            self.last_update = datetime.now(MOSCOW_TZ)
            logger.info(f"Локальный файл загружен. Найдено {len(self.products)} товаров и {len(self.shipment_dates)} дат поставок")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке локального файла: {e}")
            return False
    
    def background_ftp_update(self):
        """Фоновая загрузка данных с FTP"""
        if not self.auto_update_enabled:
            return False
            
        try:
            logger.info("🔄 Запуск фонового обновления данных с FTP...")
            success = self.download_file_from_ftp()
            if success:
                self.last_auto_update = datetime.now(MOSCOW_TZ)
                logger.info("✅ Фоновое обновление данных завершено успешно")
                return True
            else:
                logger.warning("❌ Фоновое обновление данных не удалось")
                return False
        except Exception as e:
            logger.error(f"❌ Ошибка при фоновом обновлении: {e}")
            return False
    
    def _parse_date(self, date_value):
        """Парсит дату из ячейки"""
        if not date_value:
            return None
        
        try:
            date_str = str(date_value).strip()
            date_pattern = r'(\d{1,2})\.(\d{1,2})\.(\d{4})'
            match = re.search(date_pattern, date_str)
            if match:
                day, month, year = match.groups()
                return datetime(int(year), int(month), int(day))
            return None
        except:
            return None
    
    def _parse_value(self, value):
        """Парсит значение из ячейки"""
        if value is None:
            return 0
        
        try:
            if isinstance(value, (int, float)):
                return float(value)
            
            if isinstance(value, str):
                value = value.strip()
                if value == '':
                    return 0
                if 'Более' in value:
                    return 201
                value = value.replace(' ', '').replace(',', '.')
                return float(value)
            
            return 0
        except (ValueError, TypeError):
            return 0
    
    def search_products(self, search_term):
        """Поиск товаров по артикулу"""
        if not self.products:
            return []
            
        try:
            results = []
            search_term_lower = search_term.lower()
            
            for product in self.products:
                if search_term_lower in product['name'].lower():
                    results.append(product)
            
            return results
            
        except Exception as e:
            logger.error(f"Ошибка при поиске: {e}")
            return []
    
    def format_product_info(self, product):
        """Форматирование информации о товаре для ответа с эмодзи"""
        try:
            reserve = product['reserve']
            available = product['available']
            additional_info = product['additional_info']
            
            info_suffix = f" ({additional_info})" if additional_info else ""
            
            reserve_str = "🔴 0" if reserve == 0 else f"🟢 {reserve:.3f}".rstrip('0').rstrip('.') if reserve != 201 else "🟢 Более 200"
            available_str = "🔴 0" if available == 0 else f"🟢 {available:.3f}".rstrip('0').rstrip('.') if available != 201 else "🟢 Более 200"
            
            reserve_str += info_suffix
            available_str += info_suffix
            
            product_info = f"🏷️ *{product['name']}*\n\n"
            product_info += f"🏢 *Склад Санкт-Петербург:*\n"
            product_info += f"🛡️ В резерве: {reserve_str}\n"
            product_info += f"📦 Доступно сейчас: {available_str}\n"
            
            if product['shipments']:
                sorted_shipments = sorted(
                    product['shipments'].items(),
                    key=lambda x: self._parse_date(x[0]) or datetime.max
                )
                
                product_info += f"\n🚚 *Ожидаются поступления:*\n"
                for date_display, quantity in sorted_shipments:
                    quantity_str = "🟢 Более 200" if quantity == 201 else f"🟢 {quantity:.3f}".rstrip('0').rstrip('.')
                    product_info += f"📅 {date_display}: {quantity_str}{info_suffix}\n"
            
            return product_info
                   
        except Exception as e:
            logger.error(f"Ошибка при форматировании: {e}")
            return f"❌ Ошибка при обработке товара: {product.get('name', 'Неизвестно')}"

# Глобальный экземпляр бота
stock_bot = StockBot()

# Функция для поддержания активности (только для Render.com)
async def keep_alive():
    """Периодически отправляет запросы для поддержания активности"""
    if not os.environ.get('RENDER'):
        return
        
    while True:
        try:
            # Получаем URL приложения
            app_url = f"https://{os.environ.get('RENDER_SERVICE_NAME', 'union-stock-bot')}.onrender.com"
            response = requests.get(f"{app_url}/health", timeout=10)
            logger.info(f"✅ Keep-alive запрос отправлен: {response.status_code}")
        except Exception as e:
            logger.warning(f"⚠️ Keep-alive запрос не удался: {e}")
        
        await asyncio.sleep(300)  # Каждые 5 минут

# Фоновая задача для автоматического обновления
async def auto_update_job(context: ContextTypes.DEFAULT_TYPE):
    """Фоновая задача для автоматического обновления данных"""
    success = await asyncio.to_thread(stock_bot.background_ftp_update)
    if success:
        logger.info("✅ Автоматическое обновление данных завершено")
    else:
        logger.warning("❌ Автоматическое обновление данных не удалось")

async def send_approval_request(application, user_id, username, first_name, last_name):
    """Отправка запроса на подтверждение администратору"""
    if user_id == ADMIN_ID:
        return
    
    keyboard = [
        [
            InlineKeyboardButton("✅ Подтвердить", callback_data=f"approve_{user_id}"),
            InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_{user_id}")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    user_info = f"ID: {user_id}\n"
    if username:
        user_info += f"Username: @{username}\n"
    user_info += f"Имя: {first_name or ''} {last_name or ''}".strip()
    
    message = (
        "🆕 *Новый запрос на доступ к боту*\n\n"
        f"{user_info}\n\n"
        "Выберите действие:"
    )
    
    try:
        await application.bot.send_message(
            chat_id=ADMIN_ID,
            text=message,
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        logger.info(f"✅ Запрос на доступ отправлен администратору для пользователя {user_id}")
    except Exception as e:
        logger.error(f"Не удалось отправить запрос администратору: {e}")

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    try:
        user = update.effective_user
        update_user(user.id, user.username, user.first_name, user.last_name)
        
        if user.id == ADMIN_ID:
            welcome_text = (
                "🎯 *Бот для поиска остатков товаров*\n\n"
                "🛠️ *Вы вошли как администратор*\n\n"
                "Отправьте мне артикул товара и я найду:\n"
                "• 🏢 Склад Санкт-Петербург (резерв и доступность)\n"
                "• 🚚 Ожидаемые поступления (обновляются из файла)\n\n"
                "💡 *Примеры запросов:*\n"
                "• `02-06`\n"
                "• `AR03-02`\n"
                "• `UNION 1K`\n"
                "• `Подложка`\n\n"
                "🔄 *Данные автоматически обновляются каждые 5 минут*\n"
                "⚡ *Для доступа к админ-панели отправьте /admin*"
            )
            await update.message.reply_text(welcome_text, parse_mode='Markdown')
            return
        
        if not is_user_allowed(user.id):
            user_data = get_user(user.id)
            if not user_data or not user_data.is_approved:
                pending_approvals = get_pending_approvals()
                user_pending = any(u.user_id == user.id for u in pending_approvals)
                
                if not user_pending:
                    await send_approval_request(context.application, user.id, user.username, user.first_name, user.last_name)
                
                await update.message.reply_text(
                    "⏳ *Ваш запрос на использование бота отправлен администратору.*\n\n"
                    "Ожидайте подтверждения. Вы получите уведомление, когда доступ будет предоставлен.",
                    parse_mode='Markdown'
                )
            else:
                await update.message.reply_text("❌ *Ваш аккаунт заблокирован.* Обратитесь к администратору.", parse_mode='Markdown')
            return
        
        welcome_text = (
            "🎯 *Бот для поиска остатков товаров*\n\n"
            "Отправьте мне артикул товара и я найду:\n"
            "• 🏢 Склад Санкт-Петербург (резерв и доступность)\n"
            "• 🚚 Ожидаемые поступления (обновляются из файла)\n\n"
            "💡 *Примеры запросов:*\n"
            "• `02-06`\n"
            "• `AR03-02`\n"
            "• `UNION 1K`\n"
            "• `Подложка`\n\n"
            "🔄 *Данные автоматически обновляются каждые 5 минут*"
        )
        await update.message.reply_text(welcome_text, parse_mode='Markdown')
    
    except Exception as e:
        logger.error(f"Ошибка в команде /start: {e}")
        await update.message.reply_text(
            "❌ Произошла ошибка при запуске бота. Пожалуйста, попробуйте еще раз.",
            parse_mode='Markdown'
        )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текстовых сообщений"""
    try:
        user = update.effective_user
        update_user(user.id, user.username, user.first_name, user.last_name)
        
        if user.id != ADMIN_ID:
            if not is_user_allowed(user.id):
                user_data = get_user(user.id)
                if not user_data or not user_data.is_approved:
                    await update.message.reply_text(
                        "⏳ *Ваш запрос еще не подтвержден администратором.*\n\n"
                        "Ожидайте подтверждения доступа к боту.",
                        parse_mode='Markdown'
                    )
                else:
                    await update.message.reply_text("❌ *Ваш аккаунт заблокирован.* Обратитесь к администратору.", parse_mode='Markdown')
                return
        
        user_input = update.message.text.strip()
        
        if not user_input:
            await update.message.reply_text("❌ Пожалуйста, введите артикул для поиска.")
            return
        
        status_message = await update.message.reply_text("🔍 *Поиск товаров...*", parse_mode='Markdown')
        
        try:
            products = stock_bot.search_products(user_input)
            
            if not products:
                await status_message.edit_text(f"❌ *Товары с артикулом '{user_input}' не найдены.*", parse_mode='Markdown')
                return
            
            await status_message.delete()
            
            for i, product in enumerate(products, 1):
                product_info = stock_bot.format_product_info(product)
                
                if i == len(products) and stock_bot.file_modify_time:
                    update_time = stock_bot.file_modify_time.strftime('%d.%m.%Y %H:%M')
                    product_info += f"\n\n⏰ *Данные обновлены:* {update_time}"
                
                await update.message.reply_text(product_info, parse_mode='Markdown')
                
        except Exception as e:
            logger.error(f"Ошибка при обработке запроса: {e}")
            await status_message.edit_text("❌ *Произошла ошибка при обработке запроса.*", parse_mode='Markdown')
    
    except Exception as e:
        logger.error(f"Ошибка в обработчике сообщений: {e}")
        await update.message.reply_text(
            "❌ Произошла ошибка при обработке сообщения. Пожалуйста, попробуйте еще раз.",
            parse_mode='Markdown'
        )

# Обработчик кнопок подтверждения
async def approval_button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик кнопок подтверждения пользователей"""
    try:
        query = update.callback_query
        await query.answer()
        
        if query.from_user.id != ADMIN_ID:
            await query.edit_message_text("❌ У вас нет прав для выполнения этой команды.")
            return
        
        data = query.data
        
        if data.startswith('approve_'):
            user_id = int(data.split('_')[1])
            user_data = get_user(user_id)
            
            if user_data:
                approve_user(user_id)
                log_admin_action(ADMIN_ID, "approve_user", user_id)
                
                try:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text="✅ *Ваш запрос на использование бота подтвержден!*\n\nТеперь вы можете использовать все функции бота. Отправьте /start для начала работы.",
                        parse_mode='Markdown'
                    )
                except Exception as e:
                    logger.error(f"Не удалось уведомить пользователя {user_id}: {e}")
                
                await query.edit_message_text(
                    f"✅ Пользователь {user_id} подтвержден.\nУведомление отправлено.",
                    parse_mode='Markdown'
                )
            else:
                await query.edit_message_text("❌ Пользователь не найден.")
        
        elif data.startswith('reject_'):
            user_id = int(data.split('_')[1])
            user_data = get_user(user_id)
            
            if user_data:
                block_user(user_id, "Заявка отклонена администратором")
                log_admin_action(ADMIN_ID, "reject_user", user_id)
                
                try:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text="❌ *Ваш запрос на использование бота отклонен администратором.*\n\nПо вопросам обращайтесь к администратору.",
                        parse_mode='Markdown'
                    )
                except Exception as e:
                    logger.error(f"Не удалось уведомить пользователя {user_id}: {e}")
                
                await query.edit_message_text(
                    f"❌ Пользователь {user_id} отклонен и заблокирован.\nУведомление отправлено.",
                    parse_mode='Markdown'
                )
            else:
                await query.edit_message_text("❌ Пользователь не найден.")
    
    except Exception as e:
        logger.error(f"Ошибка в обработчике кнопок подтверждения: {e}")
        try:
            await query.edit_message_text("❌ Произошла ошибка при обработке запроса.")
        except:
            pass

# АДМИН-ПАНЕЛЬ
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Панель администратора"""
    try:
        if update.effective_user.id != ADMIN_ID:
            await update.message.reply_text("❌ У вас нет прав для выполнения этой команды.")
            return
        
        keyboard = [
            [InlineKeyboardButton("📊 Статистика", callback_data="admin_stats")],
            [InlineKeyboardButton("👥 Список пользователей", callback_data="admin_users")],
            [InlineKeyboardButton("⏳ Запросы на доступ", callback_data="admin_pending")],
            [InlineKeyboardButton("🚫 Заблокированные", callback_data="admin_blocked")],
            [InlineKeyboardButton("🔄 Обновить данные", callback_data="admin_update")],
            [InlineKeyboardButton("⏰ Управление автообновлением", callback_data="admin_auto_update")],
            [InlineKeyboardButton("📋 Логи действий", callback_data="admin_logs")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "🛠️ *Панель администратора*\n\n"
            "Выберите действие:",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
    
    except Exception as e:
        logger.error(f"Ошибка в админ-панели: {e}")
        await update.message.reply_text("❌ Произошла ошибка при открытии админ-панели.")

async def admin_button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик кнопок админ-панели"""
    try:
        query = update.callback_query
        await query.answer()
        
        if query.from_user.id != ADMIN_ID:
            await query.edit_message_text("❌ У вас нет прав для выполнения этой команды.")
            return
        
        data = query.data
        
        if data == "admin_stats":
            users = get_all_users()
            total_users = len(users)
            approved_users = len([u for u in users if u.is_approved])
            pending_users = len([u for u in users if not u.is_approved and not u.is_blocked and u.user_id != ADMIN_ID])
            blocked_users = len([u for u in users if u.is_blocked])
            
            active_today = 0
            today = datetime.now(MOSCOW_TZ)
            for user in users:
                if user.last_seen:
                    if (today - user.last_seen).days < 1:
                        active_today += 1
            
            total_requests = sum(user.request_count for user in users if user.request_count)
            
            stats_text = (
                f"📊 *Статистика бота*\n\n"
                f"👥 Всего пользователей: {total_users}\n"
                f"🛠️ Администраторов: 1\n"
                f"🟢 Подтвержденных: {approved_users}\n"
                f"⏳ Ожидают подтверждения: {pending_users}\n"
                f"🚫 Заблокированных: {blocked_users}\n"
                f"📨 Активных за сегодня: {active_today}\n"
                f"📨 Всего запросов: {total_requests}\n"
                f"📦 Товаров в базе: {len(stock_bot.products)}\n"
                f"📅 Дат поставок: {len(stock_bot.shipment_dates)}\n"
                f"📡 Источник данных: {stock_bot.data_source}\n"
                f"🔄 Автообновление: {'🟢 ВКЛ' if stock_bot.auto_update_enabled else '🔴 ВЫКЛ'}\n"
                f"🌐 Хостинг: {'🟢 Render.com' if os.environ.get('RENDER') else '🔴 Локальный'}"
            )
            
            if stock_bot.last_update:
                update_time = stock_bot.last_update.strftime('%d.%m.%Y %H:%M')
                stats_text += f"\n⏰ Последнее обновление: {update_time}"
            
            await query.edit_message_text(stats_text, parse_mode='Markdown')
            
        elif data == "admin_users":
            users = get_all_users()
            if not users:
                await query.edit_message_text("👥 *Список пользователей пуст*", parse_mode='Markdown')
                return
            
            users_text = "👥 *Список пользователей:*\n\n"
            
            for user in users[:15]:
                status = "🛠️" if user.user_id == ADMIN_ID else "🚫" if user.is_blocked else "🟢" if user.is_approved else "⏳"
                username_display = f"@{user.username}" if user.username else "Без username"
                name = f"{user.first_name or ''} {user.last_name or ''}".strip()
                
                last_seen_str = user.last_seen.strftime('%d.%m.%Y %H:%M') if user.last_seen else "Неизвестно"
                
                users_text += f"{status} {user.user_id} - {username_display}\n"
                if name:
                    users_text += f"   Имя: {name}\n"
                users_text += f"   Запросов: {user.request_count}\n"
                users_text += f"   Последняя активность: {last_seen_str}\n\n"
            
            if len(users) > 15:
                users_text += f"... и еще {len(users) - 15} пользователей"
            
            keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(users_text, reply_markup=reply_markup, parse_mode='Markdown')
            
        elif data == "admin_pending":
            pending_users = get_pending_approvals()
            
            if not pending_users:
                pending_text = "⏳ *Запросов на доступ нет*"
                keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
                reply_markup = InlineKeyboardMarkup(keyboard)
                await query.edit_message_text(pending_text, reply_markup=reply_markup, parse_mode='Markdown')
                return
            
            pending_text = "⏳ *Запросы на доступ:*\n\n"
            keyboard = []
            
            for user in pending_users:
                username_display = f"@{user.username}" if user.username else "Без username"
                name = f"{user.first_name or ''} {user.last_name or ''}".strip()
                
                request_time_str = user.approval_requested.strftime('%d.%m.%Y %H:%M') if user.approval_requested else "Неизвестно"
                
                user_info = f"🆔 {user.user_id}"
                if username_display:
                    user_info += f" - {username_display}"
                if name:
                    user_info += f"\n👤 {name}"
                user_info += f"\n⏰ Запрос: {request_time_str}"
                
                pending_text += f"{user_info}\n\n"
                
                keyboard.append([
                    InlineKeyboardButton(f"✅ Подтвердить {user.user_id}", callback_data=f"approve_{user.user_id}"),
                    InlineKeyboardButton(f"❌ Отклонить {user.user_id}", callback_data=f"reject_{user.user_id}")
                ])
            
            keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_back")])
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(pending_text, reply_markup=reply_markup, parse_mode='Markdown')
            
        elif data == "admin_blocked":
            users = get_all_users()
            blocked_users = [u for u in users if u.is_blocked and u.user_id != ADMIN_ID]
            
            if not blocked_users:
                blocked_text = "🚫 *Заблокированных пользователей нет*"
                keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
                reply_markup = InlineKeyboardMarkup(keyboard)
                await query.edit_message_text(blocked_text, reply_markup=reply_markup, parse_mode='Markdown')
                return
            
            blocked_text = "🚫 *Заблокированные пользователи:*\n\n"
            keyboard = []
            
            for user in blocked_users:
                username_display = f"@{user.username}" if user.username else "Без username"
                name = f"{user.first_name or ''} {user.last_name or ''}".strip()
                
                block_info = f"Причина: {user.block_reason or 'Не указана'}\n"
                if user.block_until:
                    if datetime.now(MOSCOW_TZ) < user.block_until:
                        block_info += f"До: {user.block_until.strftime('%d.%m.%Y %H:%M')}\n"
                    else:
                        unblock_user(user.user_id)
                        continue
                
                user_info = f"🆔 {user.user_id}"
                if username_display:
                    user_info += f" - {username_display}"
                if name:
                    user_info += f"\n👤 {name}"
                user_info += f"\n{block_info}"
                
                blocked_text += f"{user_info}\n"
                
                keyboard.append([
                    InlineKeyboardButton(f"🔓 Разблокировать {user.user_id}", callback_data=f"unblock_{user.user_id}")
                ])
            
            keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="admin_back")])
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(blocked_text, reply_markup=reply_markup, parse_mode='Markdown')
            
        elif data == "admin_update":
            await query.edit_message_text("🔄 *Обновление данных...*", parse_mode='Markdown')
            
            success = stock_bot.load_data()
            
            if success:
                update_time = datetime.now(MOSCOW_TZ).strftime('%d.%m.%Y %H:%M')
                response = (
                    f"✅ *Данные успешно обновлены*\n\n"
                    f"⏰ *Время обновления:* {update_time}\n"
                    f"📡 *Источник:* {stock_bot.data_source}\n"
                    f"📊 *Товаров в базе:* {len(stock_bot.products)}\n"
                    f"📅 *Дат поставок:* {len(stock_bot.shipment_dates)}"
                )
                log_admin_action(ADMIN_ID, "update_data")
            else:
                response = "❌ *Ошибка при обновлении данных*"
            
            keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(response, reply_markup=reply_markup, parse_mode='Markdown')
            
        elif data == "admin_auto_update":
            keyboard = [
                [InlineKeyboardButton("🟢 Включить автообновление", callback_data="auto_update_on")],
                [InlineKeyboardButton("🔴 Выключить автообновление", callback_data="auto_update_off")],
                [InlineKeyboardButton("🔄 Выполнить обновление сейчас", callback_data="admin_update")],
                [InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            status_text = "🟢 ВКЛЮЧЕНО" if stock_bot.auto_update_enabled else "🔴 ВЫКЛЮЧЕНО"
            last_update = stock_bot.last_auto_update.strftime('%d.%m.%Y %H:%M') if stock_bot.last_auto_update else "Никогда"
            
            message_text = (
                f"⏰ *Управление автообновлением*\n\n"
                f"Статус: {status_text}\n"
                f"Последнее обновление: {last_update}\n"
                f"Интервал: 5 минут\n\n"
                f"Выберите действие:"
            )
            
            await query.edit_message_text(message_text, reply_markup=reply_markup, parse_mode='Markdown')
            
        elif data == "auto_update_on":
            stock_bot.auto_update_enabled = True
            log_admin_action(ADMIN_ID, "auto_update_on")
            await query.answer("✅ Автообновление включено")
            await admin_button_handler(update, context)
            
        elif data == "auto_update_off":
            stock_bot.auto_update_enabled = False
            log_admin_action(ADMIN_ID, "auto_update_off")
            await query.answer("🔴 Автообновление выключено")
            await admin_button_handler(update, context)
            
        elif data == "admin_logs":
            session = Session()
            try:
                logs = session.query(AdminLog).order_by(AdminLog.timestamp.desc()).limit(20).all()
                
                if not logs:
                    logs_text = "📋 *Логов действий нет*"
                else:
                    logs_text = "📋 *Последние действия администратора:*\n\n"
                    
                    for log in logs:
                        timestamp_str = log.timestamp.strftime('%d.%m.%Y %H:%M') if log.timestamp else str(log.timestamp)
                        
                        logs_text += f"🕐 {timestamp_str}\n"
                        logs_text += f"   Действие: {log.action}\n"
                        if log.target_user_id:
                            logs_text += f"   Пользователь: {log.target_user_id}\n"
                        if log.details:
                            logs_text += f"   Детали: {log.details}\n"
                        logs_text += "\n"
            finally:
                session.close()
            
            keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data="admin_back")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(logs_text, reply_markup=reply_markup, parse_mode='Markdown')
            
        elif data.startswith('unblock_'):
            user_id = int(data.split('_')[1])
            user_data = get_user(user_id)
            
            if user_data:
                unblock_user(user_id)
                log_admin_action(ADMIN_ID, "unblock_user", user_id)
                
                try:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text="✅ *Ваш аккаунт разблокирован администратором!*\n\nТеперь вы можете использовать бота. Отправьте /start для начала работы.",
                        parse_mode='Markdown'
                    )
                except Exception as e:
                    logger.error(f"Не удалось уведомить пользователя {user_id}: {e}")
                
                await query.edit_message_text(
                    f"✅ Пользователь {user_id} разблокирован.\nУведомление отправлено.",
                    parse_mode='Markdown'
                )
            else:
                await query.edit_message_text("❌ Пользователь не найден.")
                
        elif data == "admin_back":
            await admin_panel(update, context)
    
    except Exception as e:
        logger.error(f"Ошибка в обработчике админ-кнопок: {e}")
        try:
            await query.edit_message_text("❌ Произошла ошибка при обработке запроса.")
        except:
            pass

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ошибок"""
    logger.error(f"Exception while handling an update: {context.error}", exc_info=True)

# Веб-сервер для поддержания активности
from aiohttp import web

async def health_check(request):
    return web.Response(text="OK")

def setup_web_server():
    """Настройка веб-сервера для health checks"""
    app = web.Application()
    app.router.add_get('/health', health_check)
    return app

def main():
    """Основная функция"""
    # Создаем приложение
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Добавляем обработчики
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("admin", admin_panel))
    application.add_handler(CallbackQueryHandler(approval_button_handler, pattern="^approve_|^reject_"))
    application.add_handler(CallbackQueryHandler(admin_button_handler, pattern="^admin_|^auto_update_|^unblock_"))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_error_handler(error_handler)
    
    # Настраиваем периодическую задачу для автообновления
    job_queue = application.job_queue
    job_queue.run_repeating(auto_update_job, interval=300, first=10)
    
    # Запускаем задачу для поддержания активности (только на Render)
    if os.environ.get('RENDER'):
        loop = asyncio.get_event_loop()
        loop.create_task(keep_alive())
    
    # Предварительная загрузка данных
    print("🔄 Предварительная загрузка данных...")
    if stock_bot.load_data():
        print(f"✅ Данные загружены. Товаров: {len(stock_bot.products)}, Дат поставок: {len(stock_bot.shipment_dates)}")
        print(f"📡 Источник: {stock_bot.data_source}")
        if stock_bot.file_modify_time:
            print(f"⏰ Время обновления файла: {stock_bot.file_modify_time.strftime('%d.%m.%Y %H:%M')}")
    else:
        print("❌ Не удалось загрузить данные")
    
    # Запускаем бота
    print("🤖 Бот запущен...")
    print("🔄 Автообновление данных каждые 5 минут")
    print("⏰ Время отображается по Москве")
    print("👥 Система подтверждения пользователей активирована")
    print(f"🛠️ Администратор: {ADMIN_ID}")
    print("🌐 Хостинг: Render.com" if os.environ.get('RENDER') else "🌐 Хостинг: Локальный")
    
    # Запускаем веб-сервер для health checks (только на Render)
    if os.environ.get('RENDER'):
        port = int(os.environ.get('PORT', 8080))
        web_app = setup_web_server()
        
        async def run_web_server():
            runner = web.AppRunner(web_app)
            await runner.setup()
            site = web.TCPSite(runner, '0.0.0.0', port)
            await site.start()
            print(f"🌐 Веб-сервер запущен на порту {port}")
            
        loop = asyncio.get_event_loop()
        loop.create_task(run_web_server())
    
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()